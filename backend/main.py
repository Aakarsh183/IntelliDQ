## hi Gera ! 
## hi Shivam ! 
import os
import sys
 
os.environ['PYSPARK_PYTHON'] = sys.executable
os.environ['PYSPARK_DRIVER_PYTHON'] = sys.executable

import pandas as pd
import json
from pyspark.sql import SparkSession
from pyspark.sql.functions import *
from pyspark.sql.functions import col , monotonically_increasing_id
from pyspark.sql.functions import countDistinct
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from grok_client import GrokClient
from column_resolver import ColumnResolver

spark = SparkSession.builder.appName("DQ_ENGINE").getOrCreate()
pdf = pd.read_excel("sample_source_data.xlsx")

df = spark.createDataFrame(pdf)

dataset_columns = df.columns

rules_df = pd.read_excel("sample_rules.xlsx")

schema = "\n".join(df.columns)

client = GrokClient()
'''
column_resolver = ColumnResolver(dataset_columns)

converted_rules = []
report = []

mapped_dict = column_resolver.resolve(rules_df)
print(mapped_dict)
mapped_dict["Performance Date"] = "deposits"
mapped_dict["Period End Date"] = "perf_date"
print(mapped_dict)

def get_or_create_primary_key(df):

    for column in df.columns:
            stats = df.selectExpr(
                f"count(*) as total",
                f"count(distinct {column}) as distinct_count"
            ).collect()[0]

            if stats["total"] == stats["distinct_count"]:
                print(f"✅ Using existing primary key: {column}")
                return df, column

    print("⚠️ No primary key found. Creating new ID column...")
    df = df.withColumn("generated_id", monotonically_increasing_id())

    return df, "generated_id"

df, id_col = get_or_create_primary_key(df)

final_output = []

rule = {
        "name": "Date Sequence Validation",
        "description": "Performance date must precede period end date",
        "business_rule": "Performance Date must be less than or equal to Period End Date",
        "complexity": "medium",
        "category": "consistency"
    }
'''
# response = client.generate_pyspark_code(schema, rule, mapped_dict)
existing_rules = rules_df["business_rule"].tolist()
response = client.generate_ai_rules(schema, existing_rules)
print(response)
'''code = response["pyspark_code"]
print(code)'''

'''for _, row in rules_df.iterrows():

    rule = {
        "name": row["name"],
        "description": row["description"],
        "business_rule": row["business_rule"],
        "complexity": row["complexity"],
        "category": row["category"]
    }

    response = client.generate_pyspark_code(schema, rule, mapped_dict)

    code = response["pyspark_code"]
    print(code)
    break
    
    rule_name = rule["name"]
    
    dt_converted_rules = {"Businness_Rule":rule_name, "converted_rule":code}
    converted_rules.append(dt_converted_rules)
    print("Rule:",rule_name)
    print("Generating code", code)
    
    

rules_code_df = spark.createDataFrame(converted_rules)
rules_code_df.show()

rules_code_pd = rules_code_df.toPandas()
rules_code_pd.to_excel("rules_pyspark_mapping.xlsx", index=False)

for row in rules_code_df.collect():

    rule_name = row["Businness_Rule"]
    code = row["converted_rule"]

    print(f"\nExecuting Rule: {rule_name}")

    # execution environment
    local_vars = {
        "df": df,
        "col": col,
    }

    # execute code
    exec(code, globals(), local_vars)

    # extract results
    passed_df = local_vars["passed_df"]
    failed_df = local_vars["failed_df"]
    
    passed_ids = [row[id_col] for row in passed_df.select(id_col).collect()]
    failed_ids = [row[id_col] for row in failed_df.select(id_col).collect()]
    
    final_output.append({
                "Rule": rule["name"],
                "Passed_records": passed_ids,
                "Failed_records": failed_ids
            })
    
    dt_report = {"Business_Rule": rule_name, "Number_of_Records_passed":passed_df.count(),"Number_of_Records_failed":failed_df.count()}
    report.append(dt_report)

report_df = spark.createDataFrame(report)
report_df.show()

with open("dq_results.json", "w") as f:
    json.dump(final_output, f, indent=4)
'''

'''wb = load_workbook("rules_pyspark_mapping.xlsx")

ws = wb.active

# Apply wrap text to all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Save file
wb.save("rules_pyspark_mapping.xlsx")'''

'''passed_row_list = passed_df.select('acct_number').collect()
failed_row_list = failed_df.select('acct_number').collect()
dq_results = {}

passed_ids = [for i ]

for _, row in rules_df.itterrows():
    dq_results[row["name"]] = '''
     


    